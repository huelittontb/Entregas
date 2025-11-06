import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

ARQUIVO_BAIRROS = "bairros.xlsx"
ARQUIVO_ENTREGAS = "entregas.xlsx"
PASTA_RELATORIOS = "relatorios"
VALOR_DIARIA = 100.00


def carregar_planilha(caminho):
    if not os.path.exists(caminho):
        wb = Workbook()
        wb.active.title = "dados"
        wb.save(caminho)
    return load_workbook(caminho)


def salvar_planilha(wb, caminho):
    wb.save(caminho)


def carregar_bairros():
    if not os.path.exists(ARQUIVO_BAIRROS):
        print("Planilha de bairros não encontrada.")
        return []
    wb = load_workbook(ARQUIVO_BAIRROS)
    sheet = wb.active
    bairros = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            bairros.append({
                "cidade": row[0],
                "bairro": row[1],
                "distancia_km": row[2],
                "tempo_min": row[3],
                "valor": row[4]
            })
    return bairros


def adicionar_entrega():
    bairros = carregar_bairros()
    if not bairros:
        print("Nenhum bairro cadastrado.")
        return

    print("\nCidades disponíveis:")
    cidades = sorted(set(b["cidade"] for b in bairros))
    for i, cidade in enumerate(cidades, 1):
        print(f"{i} - {cidade}")
    try:
        escolha_cidade = int(input("Escolha a cidade: "))
        cidade_selecionada = cidades[escolha_cidade - 1]
    except (ValueError, IndexError):
        print("Opção inválida.")
        return

    bairros_cidade = [b for b in bairros if b["cidade"] == cidade_selecionada]
    print(f"\nBairros disponíveis em {cidade_selecionada}:")
    for i, b in enumerate(bairros_cidade, 1):
        print(f"{i} - {b['bairro']} (R$ {b['valor']:.2f})")

    try:
        escolha_bairro = int(input("Escolha o bairro: "))
        bairro_selecionado = bairros_cidade[escolha_bairro - 1]
    except (ValueError, IndexError):
        print("Opção inválida.")
        return

    wb = carregar_planilha(ARQUIVO_ENTREGAS)
    sheet = wb.active
    if sheet.max_row == 1:
        sheet.append(["Data", "Cidade", "Bairro", "Valor (R$)"])

    data = datetime.now().strftime("%Y-%m-%d")
    sheet.append([data, cidade_selecionada, bairro_selecionado["bairro"], bairro_selecionado["valor"]])
    salvar_planilha(wb, ARQUIVO_ENTREGAS)
    print(f"Entrega para {bairro_selecionado['bairro']} ({cidade_selecionada}) adicionada com sucesso.")


def listar_entregas():
    if not os.path.exists(ARQUIVO_ENTREGAS):
        print("Nenhuma entrega cadastrada.")
        return
    wb = load_workbook(ARQUIVO_ENTREGAS)
    sheet = wb.active

    data_hoje = datetime.now().strftime("%Y-%m-%d")
    entregas = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] == data_hoje]

    if not entregas:
        print("Nenhuma entrega registrada para hoje.")
        return

    print(f"\nEntregas do dia {data_hoje}:")
    total = 0
    for e in entregas:
        print(f"- {e[1]} / {e[2]} — R$ {e[3]:.2f}")
        total += e[3]
    print(f"\nTotal do dia (com diária de R$ {VALOR_DIARIA:.2f}): R$ {total + VALOR_DIARIA:.2f}")


def excluir_entrega():
    if not os.path.exists(ARQUIVO_ENTREGAS):
        print("Nenhuma entrega cadastrada.")
        return

    wb = load_workbook(ARQUIVO_ENTREGAS)
    sheet = wb.active
    entregas = list(sheet.iter_rows(min_row=2, values_only=True))

    if not entregas:
        print("Nenhuma entrega para excluir.")
        return

    print("\nEntregas registradas:")
    for i, e in enumerate(entregas, 1):
        print(f"{i} - {e[0]} | {e[1]} - {e[2]} | R$ {e[3]:.2f}")

    try:
        escolha = int(input("Escolha o número da entrega para excluir: "))
        if 1 <= escolha <= len(entregas):
            sheet.delete_rows(escolha + 1)
            salvar_planilha(wb, ARQUIVO_ENTREGAS)
            print("Entrega excluída com sucesso.")
        else:
            print("Opção inválida.")
    except ValueError:
        print("Entrada inválida.")


def listar_bairros():
    bairros = carregar_bairros()
    if not bairros:
        print("Nenhum bairro cadastrado.")
        return
    cidades = sorted(set(b["cidade"] for b in bairros))
    for cidade in cidades:
        print(f"\n{cidade}:")
        for b in [x for x in bairros if x["cidade"] == cidade]:
            print(f"- {b['bairro']} → R$ {b['valor']:.2f}")


def excluir_bairro():
    if not os.path.exists(ARQUIVO_BAIRROS):
        print("Planilha de bairros não encontrada.")
        return
    wb = load_workbook(ARQUIVO_BAIRROS)
    sheet = wb.active
    bairros = list(sheet.iter_rows(min_row=2, values_only=True))
    if not bairros:
        print("Nenhum bairro para excluir.")
        return
    for i, b in enumerate(bairros, 1):
        print(f"{i} - {b[0]} | {b[1]} | R$ {b[4]:.2f}")
    try:
        escolha = int(input("Escolha o número do bairro para excluir: "))
        if 1 <= escolha <= len(bairros):
            sheet.delete_rows(escolha + 1)
            salvar_planilha(wb, ARQUIVO_BAIRROS)
            print("Bairro excluído com sucesso.")
        else:
            print("Opção inválida.")
    except ValueError:
        print("Entrada inválida.")


def gerar_relatorio_diario():
    if not os.path.exists(ARQUIVO_ENTREGAS):
        print("Nenhuma entrega registrada.")
        return
    wb = load_workbook(ARQUIVO_ENTREGAS)
    sheet = wb.active

    data_hoje = datetime.now().strftime("%Y-%m-%d")
    entregas = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] == data_hoje]

    if not entregas:
        print("Nenhuma entrega para o dia de hoje.")
        return

    if not os.path.exists(PASTA_RELATORIOS):
        os.makedirs(PASTA_RELATORIOS)

    wb_rel = Workbook()
    ws = wb_rel.active
    ws.title = f"Relatório {data_hoje}"
    ws.append(["Data", "Cidade", "Bairro", "Valor (R$)"])

    total = 0
    for e in entregas:
        ws.append(e)
        total += e[3]

    ws.append(["", "", "Total com diária", total + VALOR_DIARIA])

    nome_arquivo = os.path.join(PASTA_RELATORIOS, f"relatorio_{data_hoje}.xlsx")
    wb_rel.save(nome_arquivo)
    print(f"Relatório diário gerado: {nome_arquivo}")


def menu():
    while True:
        print("\n===== SISTEMA DE CONTROLE DE ENTREGAS =====")
        print("1 - Adicionar entrega")
        print("2 - Listar entregas do dia")
        print("3 - Excluir entrega")
        print("4 - Listar bairros")
        print("5 - Excluir bairro")
        print("6 - Gerar relatório diário")
        print("0 - Sair")

        opcao = input("\nEscolha uma opção: ")

        if opcao == "1":
            adicionar_entrega()
        elif opcao == "2":
            listar_entregas()
        elif opcao == "3":
            excluir_entrega()
        elif opcao == "4":
            listar_bairros()
        elif opcao == "5":
            excluir_bairro()
        elif opcao == "6":
            gerar_relatorio_diario()
        elif opcao == "0":
            print("Saindo do sistema...")
            break
        else:
            print("Opção inválida.")


if __name__ == "__main__":
    menu()

